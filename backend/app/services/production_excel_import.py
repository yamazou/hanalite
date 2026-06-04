"""Parse production order headers from Excel (.xlsx)."""

from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.masters import Item
from app.models.production import ProductionOrder
from app.schemas.production import (
    ProductionExcelImportPreviewRow,
    ProductionOrderCreate,
    ProductionOrderUpdate,
)
from app.deps import get_tenant
from app.services.masters import MasterError, resolve_item_by_ref
from app.services.production import ProductionError


class ProductionExcelImportError(Exception):
    pass


# Normalized header -> canonical field (grid labels and legacy keys).
_CANONICAL_FIELDS: dict[str, str] = {
    "production_date": "production_date",
    "planned_date": "production_date",
    "parent_item_id": "parent_item_id",
    "item_id": "parent_item_id",
    "parent_item_cd": "parent_item_cd",
    "parent_item_code": "parent_item_cd",
    "item_cd": "parent_item_cd",
    "item_code": "parent_item_cd",
    "parent_item_nm": "parent_item_nm",
    "item_nm": "parent_item_nm",
    "item_name": "parent_item_nm",
    "planned_qty": "planned_qty",
    "plan_qty": "planned_qty",
    "lot": "lot",
    "reference_no": "reference_no",
    "reference": "reference_no",
    "notes": "notes",
}


def _normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    for ch in ".:":
        text = text.replace(ch, "")
    return text.replace(" ", "_")


def _cell_str(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _canonical_field(normalized_header: str) -> str | None:
    return _CANONICAL_FIELDS.get(normalized_header)


def _row_from_key_value_sheet(sheet) -> dict[str, Any]:
    header_cells = [c.value for c in next(sheet.iter_rows(min_row=1, max_row=1))]
    value_cells = [c.value for c in next(sheet.iter_rows(min_row=2, max_row=2))]
    if not any(header_cells) or not any(value_cells):
        raise ProductionExcelImportError(
            "Header sheet must have row 1 headers and row 2 values."
        )
    row: dict[str, Any] = {}
    for idx, raw_key in enumerate(header_cells):
        norm = _normalize_header(raw_key)
        field = _canonical_field(norm)
        if not field:
            continue
        row[field] = value_cells[idx] if idx < len(value_cells) else None
    return row


def _column_map_from_header_row(header_cells: tuple[Any, ...]) -> dict[str, int]:
    col_map: dict[str, int] = {}
    for idx, raw_key in enumerate(header_cells):
        norm = _normalize_header(raw_key)
        field = _canonical_field(norm)
        if field and field not in col_map:
            col_map[field] = idx
    return col_map


def _cell_at(row: tuple[Any, ...], col_map: dict[str, int], field: str) -> Any:
    idx = col_map.get(field)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _is_blank_data_row(row: tuple[Any, ...]) -> bool:
    return not any(v is not None and str(v).strip() != "" for v in row)


def _parse_production_date(value: Any) -> date:
    if value in (None, ""):
        raise ProductionExcelImportError("Planned Date is required.")
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        raise ProductionExcelImportError("Planned Date is required.")
    for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            continue
    raise ProductionExcelImportError("Planned Date must be YYYY-MM-DD or YYYY/MM/DD.")


def _excel_reference_lookup_key(value: Any) -> str | None:
    """Non-empty Reference No. used to match an existing order; blank always inserts."""
    return _cell_str(value)


def _order_reference_key(reference_no: str | None) -> str:
    if reference_no is None:
        return ""
    return reference_no.strip()


def _storage_reference_no(value: Any) -> str:
    """Align with UI: empty Reference No. is stored as '*'."""
    return _cell_str(value) or "*"


_PRODUCTION_STATUS_LABEL: dict[str, str] = {
    "registered": "Registered",
    "approved": "Ordered",
    "completed": "Completed",
}


def find_order_by_reference_no(db: Session, reference_key: str) -> ProductionOrder | None:
    ctx = get_tenant()
    matches = [
        row
        for row in db.scalars(
            select(ProductionOrder).where(
                ProductionOrder.co_id == ctx.co_id,
                ProductionOrder.deleted_at.is_(None),
            )
        ).all()
        if _order_reference_key(row.reference_no) == reference_key
    ]
    if len(matches) > 1:
        raise ProductionExcelImportError(
            f"Multiple production orders share reference no {reference_key!r}."
        )
    return matches[0] if matches else None


def _parse_planned_qty(value: Any) -> Decimal:
    if value in (None, ""):
        raise ProductionExcelImportError("Planned Qty is required.")
    try:
        planned_qty = Decimal(str(value))
    except (InvalidOperation, TypeError):
        raise ProductionExcelImportError("Planned Qty must be a number greater than 0.") from None
    if planned_qty <= 0:
        raise ProductionExcelImportError("Planned Qty must be a number greater than 0.")
    return planned_qty


def _row_dict_to_production_create(
    db: Session, row: dict[str, Any]
) -> tuple[ProductionOrderCreate, str | None]:
    item_id_raw = row.get("parent_item_id")
    item_cd_raw = row.get("parent_item_cd")
    item_nm_raw = row.get("parent_item_nm")
    try:
        item_id = int(item_id_raw) if item_id_raw not in (None, "") else None
    except (TypeError, ValueError):
        raise ProductionExcelImportError("Item id must be an integer.") from None

    if item_id is None and not _cell_str(item_cd_raw) and not _cell_str(item_nm_raw):
        raise ProductionExcelImportError("Item Code is required.")

    try:
        parent_item = resolve_item_by_ref(
            db,
            item_id=item_id,
            item_cd=_cell_str(item_cd_raw),
            item_nm=_cell_str(item_nm_raw),
        )
    except MasterError as e:
        raise ProductionExcelImportError(str(e)) from e

    production_date = _parse_production_date(row.get("production_date"))
    planned_qty = _parse_planned_qty(row.get("planned_qty"))
    lot = _cell_str(row.get("lot")) or "*"

    reference_lookup_key = _excel_reference_lookup_key(row.get("reference_no"))
    payload = ProductionOrderCreate(
        production_date=production_date,
        reference_no=_storage_reference_no(row.get("reference_no")),
        parent_item_id=parent_item.item_id,
        planned_qty=planned_qty,
        lot=lot,
        notes=_cell_str(row.get("notes")),
    )
    return payload, reference_lookup_key


def _create_to_header_update(payload: ProductionOrderCreate) -> ProductionOrderUpdate:
    return ProductionOrderUpdate(
        production_date=payload.production_date,
        reference_no=payload.reference_no,
        parent_item_id=payload.parent_item_id,
        planned_qty=payload.planned_qty,
        lot=payload.lot,
        notes=payload.notes,
    )


def _parse_table_rows(
    db: Session, sheet, col_map: dict[str, int]
) -> list[tuple[ProductionOrderCreate, str | None]]:
    if "production_date" not in col_map:
        raise ProductionExcelImportError("Planned Date column is required.")
    if "planned_qty" not in col_map:
        raise ProductionExcelImportError("Planned Qty column is required.")
    if "parent_item_cd" not in col_map and "parent_item_id" not in col_map and "parent_item_nm" not in col_map:
        raise ProductionExcelImportError("Item Code column is required.")

    orders: list[tuple[ProductionOrderCreate, str | None]] = []
    for row_cells in sheet.iter_rows(min_row=2, values_only=True):
        if _is_blank_data_row(row_cells):
            continue
        row_dict = {
            field: _cell_at(row_cells, col_map, field) for field in col_map
        }
        try:
            orders.append(_row_dict_to_production_create(db, row_dict))
        except ProductionExcelImportError:
            if _is_blank_data_row(row_cells):
                continue
            raise
    return orders


def _has_table_layout(col_map: dict[str, int]) -> bool:
    return (
        "production_date" in col_map
        and "planned_qty" in col_map
        and (
            "parent_item_cd" in col_map
            or "parent_item_id" in col_map
            or "parent_item_nm" in col_map
        )
    )


def parse_excel_to_production_creates(
    db: Session, raw: bytes
) -> list[tuple[ProductionOrderCreate, str | None]]:
    try:
        wb = load_workbook(BytesIO(raw), data_only=True)
    except Exception as e:  # pragma: no cover
        raise ProductionExcelImportError("Invalid .xlsx file.") from e

    if not wb.sheetnames:
        raise ProductionExcelImportError("Workbook has no sheets.")

    sheet = wb["header"] if "header" in wb.sheetnames else wb[wb.sheetnames[0]]
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    col_map = _column_map_from_header_row(header_row)

    if _has_table_layout(col_map):
        orders = _parse_table_rows(db, sheet, col_map)
        if orders:
            return orders

    kv_row = _row_from_key_value_sheet(sheet)
    return [_row_dict_to_production_create(db, kv_row)]


def parse_excel_to_production_create(db: Session, raw: bytes) -> ProductionOrderCreate:
    rows = parse_excel_to_production_creates(db, raw)
    if not rows:
        raise ProductionExcelImportError("No production orders found in Excel.")
    return rows[0][0]


def _preview_row_from_create(
    db: Session,
    *,
    excel_row: int,
    action: str,
    production_order_id: int | None,
    payload: ProductionOrderCreate,
) -> ProductionExcelImportPreviewRow:
    ctx = get_tenant()
    item = db.scalar(
        select(Item).where(
            Item.item_id == payload.parent_item_id,
            Item.co_id == ctx.co_id,
        )
    )
    if not item or item.deleted_at is not None:
        raise ProductionExcelImportError(f"Item {payload.parent_item_id} not found.")
    return ProductionExcelImportPreviewRow(
        excel_row=excel_row,
        action=action,  # type: ignore[arg-type]
        production_order_id=production_order_id,
        production_date=payload.production_date,
        reference_no=payload.reference_no,
        parent_item_id=payload.parent_item_id,
        parent_item_cd=item.item_cd,
        parent_item_nm=item.item_nm,
        planned_qty=payload.planned_qty,
        lot=payload.lot,
    )


def preview_production_orders_from_excel(
    db: Session, raw: bytes
) -> tuple[list[ProductionExcelImportPreviewRow], list[str]]:
    """Parse Excel into grid rows; DB insert/update happens on client Update."""
    rows = parse_excel_to_production_creates(db, raw)
    if not rows:
        raise ProductionExcelImportError("No production orders found in Excel.")

    preview_rows: list[ProductionExcelImportPreviewRow] = []
    row_errors: list[str] = []

    for row_index, (payload, lookup_key) in enumerate(rows):
        excel_row = row_index + 2

        if lookup_key is not None:
            try:
                existing = find_order_by_reference_no(db, lookup_key)
            except ProductionExcelImportError as e:
                row_errors.append(f"Row {excel_row}: {e}")
                continue
            if existing is not None:
                if existing.status != "registered":
                    status_label = _PRODUCTION_STATUS_LABEL.get(
                        existing.status, existing.status
                    )
                    row_errors.append(
                        f"Row {excel_row}: Reference no {lookup_key!r} cannot be updated "
                        f"(status is {status_label}; only Registered orders can be updated)."
                    )
                    continue
                try:
                    preview_rows.append(
                        _preview_row_from_create(
                            db,
                            excel_row=excel_row,
                            action="update",
                            production_order_id=int(existing.production_order_id),
                            payload=payload,
                        )
                    )
                except ProductionExcelImportError as e:
                    row_errors.append(f"Row {excel_row}: {e}")
                continue

        try:
            preview_rows.append(
                _preview_row_from_create(
                    db,
                    excel_row=excel_row,
                    action="insert",
                    production_order_id=None,
                    payload=payload,
                )
            )
        except ProductionExcelImportError as e:
            ref_label = lookup_key if lookup_key is not None else "(new)"
            row_errors.append(f"Row {excel_row}: Reference no {ref_label} — {e}")

    return preview_rows, row_errors
