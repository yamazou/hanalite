"""Parse receipt draft lines from Excel (.xlsx)."""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app.schemas.drafts import DraftCreate
from app.services.import_lines import (
    ExcelImportError,
    _match_column,
    find_header_row_index,
    parse_data_rows,
)

META_ALIASES: dict[str, list[str]] = {
    "receipt_at": ["receipt_at", "入荷日", "入荷日時", "receipt_date", "date"],
    "reference_no": ["reference_no", "参照番号", "po", "発注番号", "納品書番号"],
    "suppliers_id": ["suppliers_id", "仕入先id", "仕入先ID", "supplier_id"],
    "notes": ["notes", "備考", "note", "メモ"],
}


def _parse_meta_sheet(ws) -> dict[str, Any]:
    meta: dict[str, Any] = {}
    for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
        if not row or len(row) < 2:
            continue
        key = str(row[0]).strip() if row[0] is not None else ""
        if not key:
            continue
        val = row[1]
        for field, aliases in META_ALIASES.items():
            if _match_column(key, aliases):
                meta[field] = val
                break
    return meta


def _parse_receipt_at(value: Any) -> datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d", "%Y/%m/%d %H:%M", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    raise ExcelImportError(f"Invalid receipt_at format: {value}")


def parse_excel_to_draft_create(
    db: Session,
    file_bytes: bytes,
    *,
    receipt_at: datetime | None = None,
    suppliers_id: int | None = None,
    reference_no: str | None = None,
    notes: str | None = None,
) -> DraftCreate:
    try:
        wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        raise ExcelImportError(f"Cannot read Excel file: {e}") from e

    meta: dict[str, Any] = {}
    if "header" in wb.sheetnames or "ヘッダ" in wb.sheetnames:
        meta_sheet = wb["header"] if "header" in wb.sheetnames else wb["ヘッダ"]
        meta = _parse_meta_sheet(meta_sheet)

    lines_sheet = wb["lines"] if "lines" in wb.sheetnames else wb.active
    rows = list(lines_sheet.iter_rows(values_only=True))
    if not rows:
        raise ExcelImportError("Excel sheet is empty.")

    header_row_idx = find_header_row_index(list(rows))
    lines = parse_data_rows(db, list(rows), header_row_idx=header_row_idx)

    if not lines:
        raise ExcelImportError("No data rows found in Excel.")

    final_receipt_at = receipt_at
    if final_receipt_at is None and meta.get("receipt_at"):
        final_receipt_at = _parse_receipt_at(meta["receipt_at"])
    if final_receipt_at is None:
        final_receipt_at = datetime.now()

    final_ref = reference_no
    if final_ref is None and meta.get("reference_no"):
        final_ref = str(meta["reference_no"]).strip() or None

    final_supplier = suppliers_id
    if final_supplier is None and meta.get("suppliers_id"):
        try:
            final_supplier = int(float(str(meta["suppliers_id"])))
        except (ValueError, TypeError) as e:
            raise ExcelImportError(f"Invalid suppliers_id in Excel: {meta['suppliers_id']}") from e

    final_notes = notes
    if final_notes is None and meta.get("notes"):
        final_notes = str(meta["notes"]).strip() or None

    return DraftCreate(
        receipt_at=final_receipt_at,
        suppliers_id=final_supplier,
        reference_no=final_ref,
        notes=final_notes,
        lines=lines,
    )


def build_template_workbook() -> bytes:
    wb = Workbook()

    ws_lines = wb.active
    ws_lines.title = "lines"
    ws_lines.append(["item_id", "item_cd", "item_nm", "lot", "qty", "line_no"])
    ws_lines.append([1, "", "", "LOT-EXAMPLE-001", 100, 1])
    ws_lines.append(["", "RM-001", "Test Material A", "LOT-EXAMPLE-002", 50, 2])

    ws_header = wb.create_sheet("header")
    ws_header.append(["field", "value"])
    ws_header.append(["receipt_at", "2026-05-27 10:00:00"])
    ws_header.append(["reference_no", "PO-2026-001"])
    ws_header.append(["suppliers_id", ""])
    ws_header.append(["notes", "Excel import sample"])

    ws_help = wb.create_sheet("README")
    ws_help.append(["hanalite 入荷リスト Excel テンプレート"])
    ws_help.append([])
    ws_help.append(["lines シート: 明細（1行目=ヘッダ）"])
    ws_help.append(["  item_id / item_cd / item_nm のいずれか1つ必須"])
    ws_help.append(["  lot, qty 必須"])
    ws_help.append(["header シート: 任意（入荷日・参照番号など）"])
    ws_help.append(["アップロード画面の入荷日時は header より優先されます"])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
